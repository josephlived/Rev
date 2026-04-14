"""
SEC Ledger - Streamlit app entry point.

Run with:
    streamlit run app.py
"""
import datetime
import os
from io import BytesIO

import pandas as pd
import requests
import streamlit as st
from openpyxl.styles import Font
from streamlit.errors import StreamlitSecretNotFoundError

import edgar
import russell
from config import TARGET_FORMS


EMPTY_DISPLAY = "-"


def _parse_export_date(value):
    raw = str(value).strip()
    if not raw or raw == EMPTY_DISPLAY:
        return None
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.datetime.strptime(raw, fmt)
        except ValueError:
            continue
    parsed = pd.to_datetime(raw, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def _format_filed_date(value) -> str:
    raw = str(value).strip()
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d", "%Y%m%d"):
        try:
            parsed = datetime.datetime.strptime(raw, fmt)
            return f"{parsed.month}/{parsed.day}/{parsed.year}"
        except ValueError:
            continue
    parsed = pd.to_datetime(raw, errors="coerce")
    if pd.isna(parsed):
        return raw
    return f"{parsed.month}/{parsed.day}/{parsed.year}"


def _format_meeting_date(value) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return str(value).strip() if str(value).strip() else ""
    return f"{parsed.strftime('%B')} {parsed.day}, {parsed.year}"


def _build_display(df: pd.DataFrame) -> pd.DataFrame:
    """Reorder and rename columns for the results table."""
    out = pd.DataFrame()
    out["Date Filed"] = df["date_filed"].apply(_format_filed_date)
    out["Form Type"] = df["form_type"]
    out["Company"] = df["company_name"]
    out["Ticker"] = df["ticker"]
    out["Index"] = df["index_name"].replace("", EMPTY_DISPLAY)
    out["CIK"] = df["cik"]
    out["Meeting Type"] = df["meeting_type"].replace("", EMPTY_DISPLAY)
    out["Meeting Date"] = df["meeting_date"].apply(_format_meeting_date).replace("", EMPTY_DISPLAY)
    out["Filing"] = df["filing_url"].fillna("")
    return out


def _build_export_df(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame()
    out["Date Filed"] = df["date_filed"].apply(_parse_export_date)
    out["Form Type"] = df["form_type"]
    out["Company"] = df["company_name"]
    out["Ticker"] = df["ticker"]
    out["Index"] = df["index_name"].replace("", "")
    out["CIK"] = df["cik"]
    out["Meeting Type"] = df["meeting_type"].replace("", "")
    out["Meeting Date"] = df["meeting_date"].apply(_parse_export_date)
    out["Filing"] = df["filing_url"].fillna("")
    return out


def _build_excel_bytes(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Filings")
        worksheet = writer.sheets["Filings"]

        filing_col_idx = None
        date_col_idxs = []
        for idx, cell in enumerate(worksheet[1], start=1):
            if cell.value == "Filing":
                filing_col_idx = idx
            if cell.value in {"Date Filed", "Meeting Date"}:
                date_col_idxs.append(idx)

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value is None:
                    value = ""
                elif hasattr(cell.value, "strftime"):
                    value = f"{cell.value.month}/{cell.value.day}/{cell.value.year}"
                else:
                    value = str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 80)

        for col_idx in date_col_idxs:
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.number_format = "m/d/yyyy"

        if filing_col_idx is not None:
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=filing_col_idx)
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
                    cell.font = Font(color="0563C1", underline="single")

    output.seek(0)
    return output.getvalue()


def _show_zero_match_diagnostics(
    raw_df: pd.DataFrame,
    selected_form_df: pd.DataFrame,
    selected_forms_set: set[str],
    cik_set: set[str],
    date_label: str,
    source_label: str,
) -> None:
    st.info(
        f"No filings matched for **{date_label}**.\n\n"
        f"- **{len(raw_df):,}** total filings parsed from **{source_label}**\n"
        f"- **{len(selected_form_df):,}** matched the selected form types "
        f"({', '.join(sorted(selected_forms_set))})\n"
        f"- **0** of those had a CIK in your **{len(cik_set):,}**-company list"
    )

    if selected_form_df.empty:
        st.caption(
            "The index loaded successfully, but none of its rows matched your selected form types. "
            "That usually means the day had no matching forms, or the source file was not the daily form index you expected."
        )
        form_counts = raw_df["form_type"].value_counts().head(12).reset_index()
        if not form_counts.empty:
            form_counts.columns = ["Form Type", "Count"]
            st.dataframe(form_counts, hide_index=True, use_container_width=True)
    else:
        st.caption(
            "These filings matched the selected form types but were not in your constituent list:"
        )
        st.dataframe(
            selected_form_df[["form_type", "company_name", "cik"]].reset_index(drop=True),
            hide_index=True,
            use_container_width=True,
        )


def _get_anthropic_api_key() -> str:
    env_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if env_key:
        return env_key
    try:
        return st.secrets.get("ANTHROPIC_API_KEY", "")
    except StreamlitSecretNotFoundError:
        return ""


def _mode_uses_api(parsing_mode: str) -> bool:
    return parsing_mode == edgar.PARSING_MODE_API


def _get_cached_api_validation(api_key: str) -> str | None:
    cache = st.session_state.setdefault("anthropic_key_validation", {})
    if api_key not in cache:
        cache[api_key] = edgar.test_api_key(api_key)
    return cache[api_key]


st.set_page_config(
    page_title="SEC Ledger",
    page_icon="📋",
    layout="wide",
)

st.title("📋 SEC Ledger")
st.caption(
    "Tracks key SEC filings for US Market Coverage"
)

with st.sidebar:
    st.header("Settings")

    yesterday = datetime.date.today() - datetime.timedelta(days=1)
    col1, col2 = st.columns(2)
    start_date = col1.date_input(
        "From",
        value=yesterday,
        min_value=datetime.date(1996, 1, 1),
        max_value=datetime.date.today(),
    )
    end_date = col2.date_input(
        "To",
        value=yesterday,
        min_value=datetime.date(1996, 1, 1),
        max_value=datetime.date.today(),
    )
    if end_date < start_date:
        st.warning("'To' date must be on or after 'From' date.")
        end_date = start_date

    st.divider()

    st.subheader("Source")
    index_source = st.radio(
        "Choose how to load the EDGAR daily index",
        options=["Fetch from SEC", "Upload .idx file"],
        index=0,
        label_visibility="collapsed",
        help="Use SEC directly by default, or drag a downloaded form.YYYYMMDD.idx file here when SEC blocks access.",
    )
    idx_file = None
    if index_source == "Upload .idx file":
        idx_file = st.file_uploader(
            "Upload SEC daily index (.idx)",
            type=["idx"],
            help="Drag and drop a SEC form daily index file, such as form.20260330.idx.",
        )
        if idx_file:
            st.success(f"Using uploaded daily index: {idx_file.name}")
        else:
            st.caption("Upload a SEC daily form index file to bypass the SEC fetch step.")

    st.divider()

    st.subheader("Form Types")
    selected_forms = st.multiselect(
        "Select form types to collate",
        options=sorted(TARGET_FORMS),
        default=[],
        help="Choose which filing types to include. Deselect any you do not need.",
        placeholder="Choose options",
        label_visibility="collapsed",
    )
    if not selected_forms:
        st.warning("Select at least one form type.")

    selected_forms_set = set(selected_forms)

    st.divider()

    st.subheader("Constituent List")
    uploaded_file = st.file_uploader(
        "Upload your Excel/CSV list",
        type=["xlsx", "xls", "csv"],
        help=(
            "Upload a file with a 'CIK' or 'Ticker' column and optionally an 'Index' "
            "column (for example 'Russell 1000' or 'S&P 500'). "
            "If no file is uploaded, the app uses the iShares IWV (Russell 3000) list."
        ),
    )
    if uploaded_file:
        st.success(f"Using uploaded file: {uploaded_file.name}")
    else:
        st.caption("No file uploaded - using iShares IWV (Russell 3000).")

    st.divider()

    if not uploaded_file:
        st.subheader("Russell 3000 Cache")
        info = russell.cache_info()
        if info["exists"]:
            st.success(
                f"Cached - {info['ticker_count']} tickers\n\n"
                f"Updated: {info['updated_at']}\n\n"
                f"Age: {info['age_days']} days"
            )
        else:
            st.warning("No cache - will download on first fetch.")

        refresh_btn = st.button(
            "Refresh Russell 3000 List",
            use_container_width=True,
            help="Force re-download of the iShares IWV holdings.",
        )
    else:
        refresh_btn = False

    st.divider()

    st.subheader("Meeting Parsing")
    parsing_mode = st.selectbox(
        "Choose how DEF 14A meeting details are parsed",
        options=edgar.PARSING_MODES,
        index=edgar.PARSING_MODES.index(edgar.PARSING_MODE_REGEX),
        label_visibility="collapsed",
        help="Regex is cheapest; API uses Claude Haiku for every DEF 14A.",
    )

    api_key = ""
    api_available = False
    key_error = ""

    if _mode_uses_api(parsing_mode):
        api_key = _get_anthropic_api_key()
        if api_key:
            key_error = _get_cached_api_validation(api_key)
            if key_error:
                st.error(f"Claude API key error: {key_error}")
                api_key = ""
            else:
                api_available = True
                st.success("Claude Haiku 4.5 available")
        else:
            st.caption(
                "No API key found. Add `ANTHROPIC_API_KEY` to Streamlit secrets or environment variables."
            )

        if parsing_mode == edgar.PARSING_MODE_API and not api_available:
            st.warning("API Parsing is selected, but no valid Anthropic key is available.")
    else:
        st.caption("Regex Parsing is selected. No API calls will be made.")

    st.divider()
    fetch_btn = st.button(
        "Fetch Filings",
        use_container_width=True,
        type="primary",
        disabled=(
            not selected_forms
            or (index_source == "Upload .idx file" and idx_file is None)
            or (parsing_mode == edgar.PARSING_MODE_API and not api_available)
        ),
    )

    st.divider()
    st.caption("Data sources: SEC EDGAR and iShares IWV ETF\n\nRate-limited to <=8 req/s per SEC guidelines.")

if refresh_btn:
    with st.spinner("Downloading Russell 3000 list..."):
        status_box = st.empty()

        def _log(msg):
            status_box.info(msg)

        russell.load_russell_ciks(force_refresh=True, progress_cb=_log)
        status_box.empty()

    st.success("Russell 3000 cache refreshed.")
    st.rerun()

if fetch_btn:
    if parsing_mode == edgar.PARSING_MODE_API and not api_available:
        st.error("API Parsing requires a valid Anthropic API key before fetching can start.")
        st.stop()

    session = requests.Session()

    if uploaded_file:
        constituent_source_label = uploaded_file.name
        with st.status(f"Loading constituent list from {uploaded_file.name}...", expanded=False) as status:
            try:
                cik_set, ticker_from_cik, index_from_cik = russell.load_from_excel(uploaded_file)
                status.update(
                    label=f"Loaded {len(cik_set)} companies from {uploaded_file.name}",
                    state="complete",
                )
            except Exception as exc:
                status.update(label="Failed to load uploaded file", state="error")
                st.error(f"Could not read constituent file: {exc}")
                st.stop()
    else:
        constituent_source_label = "iShares IWV (Russell 3000)"
        with st.status("Loading Russell 3000 list...", expanded=False) as status:
            r3k_log = st.empty()

            def _r3k_log(msg):
                r3k_log.write(msg)

            try:
                cik_set, ticker_from_cik, index_from_cik = russell.load_russell_ciks(
                    progress_cb=_r3k_log
                )
                status.update(
                    label=f"Russell 3000 loaded - {len(cik_set)} companies",
                    state="complete",
                )
            except Exception as exc:
                status.update(label="Failed to load Russell 3000", state="error")
                st.error(f"Could not load Russell 3000 list: {exc}")
                st.stop()

    date_label = str(start_date) if start_date == end_date else f"{start_date} - {end_date}"
    source_label = "SEC EDGAR"
    skipped_dates: list[datetime.date] = []

    if index_source == "Upload .idx file":
        source_label = idx_file.name
        with st.status(f"Parsing uploaded daily index {idx_file.name}...", expanded=False) as status:
            try:
                raw_df = edgar.parse_daily_index_file(idx_file).drop_duplicates()
                status.update(
                    label=f"Uploaded daily index parsed - {len(raw_df):,} total filings",
                    state="complete",
                )
            except Exception as exc:
                status.update(label="Daily index parse error", state="error")
                st.error(f"Could not parse uploaded .idx file: {exc}")
                st.stop()
    else:
        dates_to_fetch = []
        current = start_date
        while current <= end_date:
            if current.weekday() < 5:
                dates_to_fetch.append(current)
            current += datetime.timedelta(days=1)

        if not dates_to_fetch:
            st.warning("The selected date range contains no weekdays. Please pick a range that includes at least one weekday.")
            st.stop()

        all_raw: list[pd.DataFrame] = []
        fetch_label = (
            f"Fetching EDGAR daily index for {date_label}..."
            if len(dates_to_fetch) == 1
            else f"Fetching EDGAR daily index ({len(dates_to_fetch)} days)..."
        )
        with st.status(fetch_label, expanded=False) as status:
            for i, day in enumerate(dates_to_fetch, 1):
                if len(dates_to_fetch) > 1:
                    status.update(label=f"Fetching {day} ({i}/{len(dates_to_fetch)})...")
                try:
                    all_raw.append(edgar.fetch_daily_index(day, session))
                except ValueError:
                    skipped_dates.append(day)
                except edgar.SecAccessError as exc:
                    status.update(label="SEC access denied", state="error")
                    st.error(
                        f"SEC denied access to the daily index for {day} (HTTP {exc.status_code}). "
                        "Retry later, or switch the sidebar source to `Upload .idx file` and drag in the SEC index manually."
                    )
                    st.caption(f"Blocked URL: {exc.url}")
                    st.stop()
                except Exception as exc:
                    status.update(label="Fetch error", state="error")
                    st.error(f"EDGAR fetch error on {day}: {exc}")
                    st.stop()

            if not all_raw:
                status.update(label="No filings found", state="error")
                st.warning(
                    f"No EDGAR filing index found for any date in the selected range ({date_label}). "
                    "This may be a holiday or market closure."
                )
                st.stop()

            raw_df = pd.concat(all_raw, ignore_index=True).drop_duplicates()
            complete_label = f"Daily index fetched - {len(raw_df):,} total filings"
            if skipped_dates:
                complete_label += f" ({len(skipped_dates)} date(s) skipped - no index)"
            status.update(label=complete_label, state="complete")

    if "form_type_normalized" not in raw_df.columns:
        raw_df["form_type_normalized"] = raw_df["form_type"].map(edgar.normalize_form_type)

    normalized_selected_forms = {edgar.normalize_form_type(form) for form in selected_forms_set}
    selected_form_df = raw_df[raw_df["form_type_normalized"].isin(normalized_selected_forms)].copy()

    filtered_df = edgar.filter_filings(raw_df, cik_set, target_forms=selected_forms_set)
    filtered_df = edgar.enrich_with_ticker(filtered_df, ticker_from_cik, index_from_cik)

    if filtered_df.empty:
        _show_zero_match_diagnostics(
            raw_df=raw_df,
            selected_form_df=selected_form_df,
            selected_forms_set=selected_forms_set,
            cik_set=cik_set,
            date_label=date_label,
            source_label=source_label,
        )
        st.caption(f"Constituent source: {constituent_source_label}")
        st.stop()

    def14a_count = (
        filtered_df["form_type_normalized"] == edgar.normalize_form_type("DEF 14A")
    ).sum()

    if def14a_count > 0:
        progress_bar = st.progress(0, text="Parsing DEF 14A filings...")

        def _def14a_progress(msg, current, total):
            pct = int(current / total * 100)
            progress_bar.progress(pct, text=f"{msg} ({current}/{total})")

        filtered_df = edgar.parse_def14a_filings(
            filtered_df,
            session,
            progress_cb=_def14a_progress,
            parsing_mode=parsing_mode,
            api_key=api_key or None,
        )
        progress_bar.empty()
        if _mode_uses_api(parsing_mode):
            errors = filtered_df.loc[
                filtered_df["claude_error"].str.len() > 0, "claude_error"
            ]
            if not errors.empty:
                st.warning(
                    f"Claude API failed on {len(errors)} filing(s); meeting date left blank. "
                    f"First error: {errors.iloc[0]}"
                )

        if "parsing_method" in filtered_df.columns:
            api_used = (filtered_df["parsing_method"] == "api").sum()
            regex_used = (filtered_df["parsing_method"] == "regex").sum()
            st.caption(
                f"Meeting parsing mode: {parsing_mode}. Regex handled {regex_used} filing(s); API handled {api_used} filing(s)."
            )
    else:
        filtered_df["meeting_type"] = ""
        filtered_df["meeting_date"] = ""

    filing_progress_bar = st.progress(0, text="Resolving filing links...")

    def _filing_progress(msg, current, total):
        pct = int(current / total * 100) if total else 100
        filing_progress_bar.progress(pct, text=f"{msg} ({current}/{total})")

    filtered_df = edgar.enrich_with_filing_url(
        filtered_df,
        session,
        progress_cb=_filing_progress,
    )
    filing_progress_bar.empty()

    display_df = _build_display(filtered_df)

    form_counts = filtered_df["form_type"].value_counts().to_dict()
    cols = st.columns(len(form_counts) + 1)
    cols[0].metric("Total Filings", len(filtered_df))
    for i, (form, count) in enumerate(sorted(form_counts.items()), start=1):
        cols[i].metric(form, count)

    st.divider()
    st.subheader(f"Results for {date_label}")
    st.caption(f"Daily index source: {source_label}")
    st.dataframe(display_df, use_container_width=True, hide_index=True)

    if start_date == end_date:
        xlsx_filename = f"sec_filings_{start_date}.xlsx"
    else:
        xlsx_filename = f"sec_filings_{start_date}_to_{end_date}.xlsx"

    export_df = _build_export_df(filtered_df)
    xlsx_bytes = _build_excel_bytes(export_df)
    st.download_button(
        label="Export",
        data=xlsx_bytes,
        file_name=xlsx_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
